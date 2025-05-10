import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import load_workbook


def read_linear_regression_data(file_path):
    try:
        # Загрузка Excel-файла с вычислением формул
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb['Линейная регрессия']  # Указываем лист

        # Чтение данных из конкретных ячеек
        a = sheet['B8'].value
        b = sheet['C8'].value
        h = sheet['D8'].value
        k = sheet['D4'].value

        # Чтение массива x из ячеек B16:B(16+k-1)
        sample_x = []
        for i in range(k):
            cell = sheet[f'B{16 + i}']
            if cell.value is not None:
                try:
                    sample_x.append(float(cell.value))
                except ValueError:
                    print(f"Ошибка: значение в ячейке B{16 + i} не является числом.")
                    return None, None, None, None, None, None

        # Чтение массива z из ячеек D16:D(16+k-1)
        sample_z = []
        for i in range(k):
            cell = sheet[f'D{16 + i}']
            if cell.value is not None:
                try:
                    sample_z.append(float(cell.value))
                except ValueError:
                    print(f"Ошибка: значение в ячейке D{16 + i} не является числом.")
                    return None, None, None, None, None, None

        # Проверяем, что выборки не пустые
        if len(sample_x) == 0 or len(sample_z) == 0:
            raise ValueError("Выборка пуста. Проверьте диапазон ячеек B16:B(16+k-1) и D16:D(16+k-1).")

        return a, b, h, k, sample_x, sample_z
    except Exception as e:
        print(f"Ошибка при чтении данных из Excel: {e}")
        return None, None, None, None, None, None


def read_nonlinear_regression_data(file_path):
    try:
        # Загрузка Excel-файла с вычислением формул
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb['Нелин. парн.регрессия']  # Указываем лист

        # Чтение данных из конкретных ячеек
        a = sheet['B8'].value
        b = sheet['C8'].value
        c = sheet['D8'].value
        h = sheet['E8'].value
        k = sheet['D4'].value

        # Чтение массива x из ячеек B16:B(16+k-1)
        sample_x = []
        for i in range(k):
            cell = sheet[f'B{16 + i}']
            if cell.value is not None:
                try:
                    sample_x.append(float(cell.value))
                except ValueError:
                    print(f"Ошибка: значение в ячейке B{16 + i} не является числом.")
                    return None, None, None, None, None, None, None

        # Чтение массива z из ячеек E16:E(16+k-1)
        sample_z = []
        for i in range(k):
            cell = sheet[f'E{16 + i}']
            if cell.value is not None:
                try:
                    sample_z.append(float(cell.value))
                except ValueError:
                    print(f"Ошибка: значение в ячейке E{16 + i} не является числом.")
                    return None, None, None, None, None, None, None

        # Проверка, что данные не пустые и являются числами
        if None in (a, b, c, h, k):
            raise ValueError("Одна из ячеек B8, C8, D8, E8, D4 пустая.")
        if not all(isinstance(val, (int, float)) for val in (a, b, c, h, k)):
            raise ValueError("Одна из ячеек B8, C8, D8, E8, D4 содержит нечисловые данные.")

        return a, b, c, h, k, sample_x, sample_z
    except Exception as e:
        print(f"Ошибка при чтении данных из Excel: {e}")
        return None, None, None, None, None, None, None


def read_multiple_nonlinear_regression_data(file_path):
    try:
        # Загрузка Excel-файла с вычислением формул
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb['Нелин. множ. регрессия']  # Указываем лист

        # Чтение данных из конкретных ячеек
        a0 = sheet['C8'].value
        a1 = sheet['D8'].value
        a2 = sheet['E8'].value
        a11 = sheet['F8'].value
        a22 = sheet['G8'].value
        a12 = sheet['H8'].value
        h = sheet['I8'].value
        k = sheet['D4'].value

        # Чтение массива x1 из ячеек B16:B(16+k-1)
        sample_x1 = []
        for i in range(k):
            cell = sheet[f'B{16 + i}']
            if cell.value is not None:
                try:
                    sample_x1.append(float(cell.value))
                except ValueError:
                    print(f"Ошибка: значение в ячейке B{16 + i} не является числом.")
                    return None, None, None, None, None, None, None, None, None, None

        # Чтение массива x2 из ячеек C16:C(16+k-1)
        sample_x2 = []
        for i in range(k):
            cell = sheet[f'C{16 + i}']
            if cell.value is not None:
                try:
                    sample_x2.append(float(cell.value))
                except ValueError:
                    print(f"Ошибка: значение в ячейке C{16 + i} не является числом.")
                    return None, None, None, None, None, None, None, None, None, None

        # Чтение массива z из ячеек H16:H(16+k-1)
        sample_z = []
        for i in range(k):
            cell = sheet[f'H{16 + i}']
            if cell.value is not None:
                try:
                    sample_z.append(float(cell.value))
                except ValueError:
                    print(f"Ошибка: значение в ячейке H{16 + i} не является числом.")
                    return None, None, None, None, None, None, None, None, None, None

        # Проверка, что данные не пустые и являются числами
        if None in (a0, a1, a2, a11, a22, a12, h, k):
            raise ValueError("Одна из ячеек B8, C8, D8, E8, F8, G8, H8, D4 пустая.")
        if not all(isinstance(val, (int, float)) for val in (a0, a1, a2, a11, a22, a12, h, k)):
            raise ValueError("Одна из ячеек B8, C8, D8, E8, F8, G8, H8, D4 содержит нечисловые данные.")

        return a0, a1, a2, a11, a22, a12, h, k, sample_x1, sample_x2, sample_z
    except Exception as e:
        print(f"Ошибка при чтении данных из Excel: {e}")
        return None, None, None, None, None, None, None, None, None, None


def input_linear_data_manually():
    try:
        a = float(input("Введите a для линейной регрессии: "))
        b = float(input("Введите b для линейной регрессии: "))
        h = float(input("Введите h для линейной регрессии: "))
        k = int(input("Введите k для линейной регрессии: "))
        return a, b, h, k
    except ValueError:
        print("Ошибка: введены некорректные данные. Убедитесь, что вводите числа.")
        return None, None, None, None


def input_nonlinear_data_manually():
    try:
        a = float(input("Введите a для нелинейной регрессии: "))
        b = float(input("Введите b для нелинейной регрессии: "))
        c = float(input("Введите c для нелинейной регрессии: "))
        h = float(input("Введите h для нелинейной регрессии: "))
        k = int(input("Введите k для нелинейной регрессии: "))
        return a, b, c, h, k
    except ValueError:
        print("Ошибка: введены некорректные данные. Убедитесь, что вводите числа.")
        return None, None, None, None, None


def input_multiple_nonlinear_data_manually():
    try:
        a0 = float(input("Введите a0 для нелинейной множественной регрессии: "))
        a1 = float(input("Введите a1 для нелинейной множественной регрессии: "))
        a2 = float(input("Введите a2 для нелинейной множественной регрессии: "))
        a11 = float(input("Введите a11 для нелинейной множественной регрессии: "))
        a22 = float(input("Введите a22 для нелинейной множественной регрессии: "))
        a12 = float(input("Введите a12 для нелинейной множественной регрессии: "))
        h = float(input("Введите h для нелинейной множественной регрессии: "))
        k = int(input("Введите k для нелинейной множественной регрессии: "))
        return a0, a1, a2, a11, a22, a12, h, k
    except ValueError:
        print("Ошибка: введены некорректные данные. Убедитесь, что вводите числа.")
        return None, None, None, None, None, None, None, None


def generate_linear_data(a, b, h, k):
    try:
        x = np.arange(1, k + 1)  # Массив x от 1 до k
        z = np.random.normal(0, 1, k)  # Массив z с нормальным распределением
        y = a + b * x + h * z  # Массив y

        return x, y, z
    except Exception as e:
        print(f"Ошибка при генерации данных: {e}")
        return None, None, None


def generate_nonlinear_data(a, b, c, h, k):
    try:
        x = np.arange(1, k + 1)  # Массив x от 1 до k
        z = np.random.normal(0, 1, k)  # Массив z с нормальным распределением
        y = a + b * x + c * x ** 2 + h * z  # Массив y

        return x, y, z
    except Exception as e:
        print(f"Ошибка при генерации данных: {e}")
        return None, None, None


def generate_multiple_nonlinear_data(a0, a1, a2, a11, a22, a12, h, k):
    try:
        x1 = np.arange(0, k / 10, 0.1)  # Массив x1 от 0 до (k/10) - 0.1 с шагом 0.1
        x2 = np.arange(0, k / 10, 0.1)  # Массив x2 от 0 до (k/10) - 0.1 с шагом 0.1
        z = np.random.normal(0, 1, len(x1))  # Массив z с нормальным распределением
        y = a0 + a1 * x1 + a2 * x2 + a11 * x1 ** 2 + a22 * x2 ** 2 + a12 * x1 * x2 + h * z  # Массив y

        return x1, x2, y, z
    except Exception as e:
        print(f"Ошибка при генерации данных: {e}")
        return None, None, None, None


def fit_linear_regression(x, y):
    # Вычисление коэффициентов a' и b' для уравнения y = a' + b'*x
    A = np.vstack([x, np.ones(len(x))]).T
    b_fit, a_fit = np.linalg.lstsq(A, y, rcond=None)[0]
    return a_fit, b_fit


def fit_nonlinear_regression(x, y):
    # Вычисление коэффициентов a', b', c' для уравнения y = a' + b'*x + c'*x^2
    A = np.vstack([x ** 2, x, np.ones(len(x))]).T
    c_fit, b_fit, a_fit = np.linalg.lstsq(A, y, rcond=None)[0]
    return a_fit, b_fit, c_fit


def fit_multiple_nonlinear_regression(x1, x2, y):
    # Создание матрицы A для уравнения y = a0 + a1*x1 + a2*x2 + a11*x1^2 + a22*x2^2 + a12*x1*x2
    A = np.vstack([x1 ** 2, x2 ** 2, x1 * x2, x1, x2, np.ones(len(x1))]).T
    # Решение системы уравнений
    coefficients, _, _, _ = np.linalg.lstsq(A, y, rcond=None)
    a11_fit, a22_fit, a12_fit, a1_fit, a2_fit, a0_fit = coefficients
    return a0_fit, a1_fit, a2_fit, a11_fit, a22_fit, a12_fit


def plot_linear_results(x, y, a_fit, b_fit):
    # Построение графика исходных данных
    plt.scatter(x, y, label='Исходные данные (x, y)', color='blue')

    # Построение линии тренда
    y_trend = a_fit + b_fit * x
    plt.plot(x, y_trend, color='red', label=f'Линия тренда: y = {a_fit:.2f} + {b_fit:.2f}x')

    # Настройка графика
    plt.xlabel('x')
    plt.ylabel('y')
    plt.title('Линейная парная регрессия и линия тренда')
    plt.legend()
    plt.grid(True)
    plt.show()


def plot_nonlinear_results(x, y, a_fit, b_fit, c_fit):
    # Построение графика исходных данных
    plt.scatter(x, y, label='Исходные данные (x, y)', color='blue')

    # Построение нелинейной линии тренда
    y_trend_nonlinear = a_fit + b_fit * x + c_fit * x ** 2
    plt.plot(x, y_trend_nonlinear, color='green',
             label=f'Полиномная линия тренда: y = {a_fit:.2f} + {b_fit:.2f}x + {c_fit:.2f}x^2')

    # Построение линейной линии тренда
    a_fit_linear, b_fit_linear = fit_linear_regression(x, y)
    y_trend_linear = a_fit_linear + b_fit_linear * x
    plt.plot(x, y_trend_linear, color='red',
             label=f'Линейная линия тренда: y = {a_fit_linear:.2f} + {b_fit_linear:.2f}x')

    # Настройка графика
    plt.xlabel('x')
    plt.ylabel('y')
    plt.title('Нелинейная парная регрессия и линии тренда')
    plt.legend()
    plt.grid(True)
    plt.show()


def plot_multiple_nonlinear_results(x1, x2, y, a0_fit, a1_fit, a2_fit, a11_fit, a22_fit, a12_fit):
    # Обычный график y(x)
    x = np.arange(1, len(y) + 1)  # Массив x от 1 до k
    y_trend_nonlinear = (
            a0_fit
            + a1_fit * x1
            + a2_fit * x2
            + a11_fit * x1 ** 2
            + a22_fit * x2 ** 2
            + a12_fit * x1 * x2
    )

    # Линейная линия тренда
    a_fit_linear, b_fit_linear = fit_linear_regression(x, y)
    y_trend_linear = a_fit_linear + b_fit_linear * x

    plt.figure(figsize=(10, 5))
    plt.scatter(x, y, label='Исходные данные', color='blue')
    plt.plot(x, y_trend_nonlinear, color='red', label=f'Полиномная линия тренда')
    plt.plot(x, y_trend_linear, color='green',
             label=f'Линейная линия тренда: y = {a_fit_linear:.2f} + {b_fit_linear:.2f}x')
    plt.xlabel('x')
    plt.ylabel('y')
    plt.title('График y(x) для нелинейной множественной регрессии')
    plt.legend()
    plt.grid(True)
    plt.show()

    # 3D-график поверхности регрессии
    fig = plt.figure(figsize=(10, 7))
    ax = fig.add_subplot(111, projection='3d')

    # Создание сетки для x1 и x2
    x1_range = np.linspace(min(x1), max(x1), 100)
    x2_range = np.linspace(min(x2), max(x2), 100)
    x1_grid, x2_grid = np.meshgrid(x1_range, x2_range)

    # Вычисление предсказанных значений для 3D-графика
    y_trend_3d = (
            a0_fit
            + a1_fit * x1_grid
            + a2_fit * x2_grid
            + a11_fit * x1_grid ** 2
            + a22_fit * x2_grid ** 2
            + a12_fit * x1_grid * x2_grid
    )

    # Построение поверхности
    surf = ax.plot_surface(x1_grid, x2_grid, y_trend_3d, cmap='viridis', alpha=0.7, label='Поверхность регрессии')
    ax.scatter(x1, x2, y, color='blue', label='Исходные данные')

    # Добавление цветовой легенды
    fig.colorbar(surf, ax=ax, shrink=0.5, aspect=5)

    # Настройка графика
    ax.set_xlabel('x1')
    ax.set_ylabel('x2')
    ax.set_zlabel('y')
    ax.set_title('3D-график поверхности регрессии')
    plt.legend()
    plt.show()


def run_test(file_path):
    # Линейная регрессия
    a_linear, b_linear, h_linear, k_linear, sample_x, sample_z = read_linear_regression_data(file_path)
    if None in (a_linear, b_linear, h_linear, k_linear, sample_x, sample_z):
        print("Не удалось загрузить данные для линейной регрессии из Excel. Проверьте файл и ячейки.")
    else:
        print(
            f"Считанные данные для линейной регрессии: a = {a_linear}, b = {b_linear}, h = {h_linear}, k = {k_linear}")
        x_linear = np.array(sample_x)
        z_linear = np.array(sample_z)
        y_linear = a_linear + b_linear * x_linear + h_linear * z_linear

        # Подбор коэффициентов линии тренда
        a_fit_linear, b_fit_linear = fit_linear_regression(x_linear, y_linear)
        print(f"Коэффициенты линии тренда (линейная регрессия): a' = {a_fit_linear:.2f}, b' = {b_fit_linear:.2f}")

        # Построение графика
        plot_linear_results(x_linear, y_linear, a_fit_linear, b_fit_linear)

    # Нелинейная регрессия
    a_nonlinear, b_nonlinear, c_nonlinear, h_nonlinear, k_nonlinear, sample_x, sample_z = read_nonlinear_regression_data(
        file_path)
    if None in (a_nonlinear, b_nonlinear, c_nonlinear, h_nonlinear, k_nonlinear, sample_x, sample_z):
        print("Не удалось загрузить данные для нелинейной регрессии из Excel. Проверьте файл и ячейки.")
    else:
        print(
            f"Считанные данные для нелинейной регрессии: a = {a_nonlinear}, b = {b_nonlinear}, c = {c_nonlinear}, h = {h_nonlinear}, k = {k_nonlinear}")
        x_nonlinear = np.array(sample_x)
        z_nonlinear = np.array(sample_z)
        y_nonlinear = a_nonlinear + b_nonlinear * x_nonlinear + c_nonlinear * x_nonlinear ** 2 + h_nonlinear * z_nonlinear

        # Подбор коэффициентов линии тренда
        a_fit_nonlinear, b_fit_nonlinear, c_fit_nonlinear = fit_nonlinear_regression(x_nonlinear, y_nonlinear)
        print(
            f"Коэффициенты линии тренда (нелинейная регрессия): a' = {a_fit_nonlinear:.2f}, b' = {b_fit_nonlinear:.2f}, c' = {c_fit_nonlinear:.2f}")

        # Построение графика
        plot_nonlinear_results(x_nonlinear, y_nonlinear, a_fit_nonlinear, b_fit_nonlinear, c_fit_nonlinear)

    # Нелинейная множественная регрессия
    a0, a1, a2, a11, a22, a12, h, k, sample_x1, sample_x2, sample_z = read_multiple_nonlinear_regression_data(file_path)
    if None in (a0, a1, a2, a11, a22, a12, h, k, sample_x1, sample_x2, sample_z):
        print("Не удалось загрузить данные для нелинейной множественной регрессии из Excel. Проверьте файл и ячейки.")
    else:
        print(
            f"Считанные данные для нелинейной множественной регрессии: a0 = {a0}, a1 = {a1}, a2 = {a2}, a11 = {a11}, a22 = {a22}, a12 = {a12}, h = {h}, k = {k}")
        x1 = np.array(sample_x1)
        x2 = np.array(sample_x2)
        z = np.array(sample_z)
        y = a0 + a1 * x1 + a2 * x2 + a11 * x1 ** 2 + a22 * x2 ** 2 + a12 * x1 * x2 + h * z

        # Подбор коэффициентов линии тренда
        a0_fit, a1_fit, a2_fit, a11_fit, a22_fit, a12_fit = fit_multiple_nonlinear_regression(x1, x2, y)
        print(
            f"Коэффициенты линии тренда (нелинейная множественная регрессия): a0' = {a0_fit:.2f}, a1' = {a1_fit:.2f}, a2' = {a2_fit:.2f}, a11' = {a11_fit:.2f}, a22' = {a22_fit:.2f}, a12' = {a12_fit:.2f}")

        # Построение графиков
        plot_multiple_nonlinear_results(x1, x2, y, a0_fit, a1_fit, a2_fit, a11_fit, a22_fit, a12_fit)


def run_work():
    # Линейная регрессия
    a_linear, b_linear, h_linear, k_linear = input_linear_data_manually()
    if None in (a_linear, b_linear, h_linear, k_linear):
        print("Ошибка: введены некорректные данные для линейной регрессии.")
    else:
        print(
            f"Введенные данные для линейной регрессии: a = {a_linear}, b = {b_linear}, h = {h_linear}, k = {k_linear}")
        x_linear, y_linear, z_linear = generate_linear_data(a_linear, b_linear, h_linear, k_linear)
        if x_linear is None or y_linear is None or z_linear is None:
            print("Ошибка при генерации данных для линейной регрессии.")
        else:
            a_fit_linear, b_fit_linear = fit_linear_regression(x_linear, y_linear)
            print(f"Коэффициенты линии тренда (линейная регрессия): a' = {a_fit_linear:.2f}, b' = {b_fit_linear:.2f}")
            plot_linear_results(x_linear, y_linear, a_fit_linear, b_fit_linear)

    # Нелинейная регрессия
    a_nonlinear, b_nonlinear, c_nonlinear, h_nonlinear, k_nonlinear = input_nonlinear_data_manually()
    if None in (a_nonlinear, b_nonlinear, c_nonlinear, h_nonlinear, k_nonlinear):
        print("Ошибка: введены некорректные данные для нелинейной регрессии.")
    else:
        print(
            f"Введенные данные для нелинейной регрессии: a = {a_nonlinear}, b = {b_nonlinear}, c = {c_nonlinear}, h = {h_nonlinear}, k = {k_nonlinear}")
        x_nonlinear, y_nonlinear, z_nonlinear = generate_nonlinear_data(a_nonlinear, b_nonlinear, c_nonlinear,
                                                                        h_nonlinear, k_nonlinear)
        if x_nonlinear is None or y_nonlinear is None or z_nonlinear is None:
            print("Ошибка при генерации данных для нелинейной регрессии.")
        else:
            a_fit_nonlinear, b_fit_nonlinear, c_fit_nonlinear = fit_nonlinear_regression(x_nonlinear, y_nonlinear)
            print(
                f"Коэффициенты линии тренда (нелинейная регрессия): a' = {a_fit_nonlinear:.2f}, b' = {b_fit_nonlinear:.2f}, c' = {c_fit_nonlinear:.2f}")
            plot_nonlinear_results(x_nonlinear, y_nonlinear, a_fit_nonlinear, b_fit_nonlinear, c_fit_nonlinear)

    # Нелинейная множественная регрессия
    a0, a1, a2, a11, a22, a12, h, k = input_multiple_nonlinear_data_manually()
    if None in (a0, a1, a2, a11, a22, a12, h, k):
        print("Ошибка: введены некорректные данные для нелинейной множественной регрессии.")
    else:
        print(
            f"Введенные данные для нелинейной множественной регрессии: a0 = {a0}, a1 = {a1}, a2 = {a2}, a11 = {a11}, a22 = {a22}, a12 = {a12}, h = {h}, k = {k}")
        x1, x2, y, z = generate_multiple_nonlinear_data(a0, a1, a2, a11, a22, a12, h, k)
        if x1 is None or x2 is None or y is None or z is None:
            print("Ошибка при генерации данных для нелинейной множественной регрессии.")
        else:
            a0_fit, a1_fit, a2_fit, a11_fit, a22_fit, a12_fit = fit_multiple_nonlinear_regression(x1, x2, y)
            print(
                f"Коэффициенты линии тренда (нелинейная множественная регрессия): a0' = {a0_fit:.2f}, a1' = {a1_fit:.2f}, a2' = {a2_fit:.2f}, a11' = {a11_fit:.2f}, a22' = {a22_fit:.2f}, a12' = {a12_fit:.2f}")
            plot_multiple_nonlinear_results(x1, x2, y, a0_fit, a1_fit, a2_fit, a11_fit, a22_fit, a12_fit)


if __name__ == "__main__":
    # Путь к Excel-файлу
    file_path = 'AsOr_Zadanie_2_1_2_2_2_3.xlsx'

    # Выбор программы
    choice = input("test/work\n").strip().lower()

    if choice == "test":
        print("Запуск тестовой программы...")
        run_test(file_path)
    elif choice == "work":
        print("Запуск рабочей программы...")
        run_work()
    else:
        print("Ошибка: выберите test или work.")
